#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_excel_data.py — 三安光电预算系统 · Excel 客户真实主数据全量导入
来源：docs/product/ 三个 xlsx
  - 各中心费控系统2026年预算数据收集模板：公司对照表(42家)/事业部生产单元对照表/
    一级部门对照表(42)/费控科目明细表(81经济事项+17归口部门)
  - 2.预算科目及逻辑-总经办表单：预算逻辑(控制逻辑文字)
  - 3.总经办预算汇报-汇总表：集团汇总历史（本版未逐公司匹配，金额生成式）

说明：Excel 为模板/字典，不含逐公司逐科目真实预算金额，因此
  - organization / economic_event（科目+会计编码+归口部门+控制逻辑）为真实全量导入
  - unit_budget 金额 = 经济事项基线 × 公司规模系数（生成式，关联真实组织与事项）
    （客户填好真实表单后可重跑本脚本一键覆盖）
写入 server/economic_event.db。
"""
import openpyxl, sqlite3, os, hashlib, json

BASE = "/Users/yangjackson/AIProjects/BadgetManager"
XLSX = os.path.join(BASE, "docs/product/")
DB = os.path.join(BASE, "server/economic_event.db")

MAIN = "各中心费控系统2026年预算数据收集模板-财务部费控系统表单.xlsx"
LOGIC = "2.预算科目及逻辑-总经办表单.xlsx"


def load(fname):
    return openpyxl.load_workbook(os.path.join(XLSX, fname), data_only=True, read_only=True)


# ---------------- 解析 Excel ----------------
wb = load(MAIN)

# 公司对照表
ws = wb["公司代码公司名称对照表"]
companies = []
for r in ws.iter_rows(values_only=True):
    if not r or not r[0]:
        continue
    code = str(r[0]).strip()
    if code == "公司代码":
        continue
    name = str(r[1]).strip() if r[1] else code
    companies.append((code, name))

# 事业部 + 生产单元
ws = wb["事业部生产单元对照表"]
business = {}
units = []
for r in ws.iter_rows(values_only=True):
    if not r or not r[0]:
        continue
    bc = str(r[0]).strip()
    bn = str(r[1]).strip() if r[1] else bc
    if bc == "事业部组织编码":
        continue
    business[bc] = bn
    if len(r) > 4 and r[4]:
        uc = str(r[4]).strip()
        un = str(r[5]).strip() if len(r) > 5 and r[5] else uc
        units.append((uc, un, bc))

# 一级部门
ws = wb["一级部门对照表"]
depts = []
for r in ws.iter_rows(values_only=True):
    if not r or len(r) < 3 or not r[1]:
        continue
    dc = str(r[1]).strip()
    if not dc.isdigit():
        continue
    dn = str(r[2]).strip() if r[2] else dc
    depts.append((dc, dn))

# 费控科目明细表 -> 经济事项 + 归口部门
ws = wb["费控系统预算科目明细表"]
events = {}
centers = set()
for r in ws.iter_rows(values_only=True):
    if not r or not r[4]:
        continue
    cat = str(r[4]).strip()
    if cat == "预算科目名称":
        continue
    acct = str(r[0]).strip() if r[0] else ""
    center = str(r[5]).strip() if r[5] else ""
    # 类别：优先「所属分类名称」(col[2])，缺则「费用类型名称」(col[3]) —— 均为客户真实分类
    cat_cls = str(r[2]).strip() if r[2] else (str(r[3]).strip() if r[3] else "")
    if cat not in events:
        events[cat] = {"acct": acct, "center": center, "cat": cat_cls, "method": ""}
    if center:
        centers.add(center)
wb.close()

# 预算逻辑（控制逻辑文字）
wb = load(LOGIC)
ws = wb["1.预算逻辑"]
logic_map = {}
for r in ws.iter_rows(values_only=True):
    if not r or len(r) < 6 or not r[1]:
        continue
    cat = str(r[1]).strip()
    if cat == "预算科目名称":
        continue
    logic = str(r[5]).strip() if r[5] else ""
    if cat not in logic_map:
        logic_map[cat] = logic
wb.close()


# ---------------- 确定性辅助 ----------------
def det(seed, lo, hi):
    h = int(hashlib.md5(seed.encode("utf-8")).hexdigest(), 16)
    return lo + (h % (hi - lo))


BASE_RATIO = [0.07, 0.06, 0.08, 0.07, 0.08, 0.09, 0.08, 0.09, 0.1, 0.09, 0.1, 0.09]


def decompose(total):
    base = [round((total or 0) * p) for p in BASE_RATIO]
    s = sum(base)
    base[11] += (total or 0) - s
    return base


def baseline(method, ly):
    ly = ly or 0
    return {
        "down5": round(ly * 0.95), "canteen": round(ly * 0.97), "dorm": round(ly * 0.90),
        "revenue": round(ly * 0.98), "green": round(ly * 0.92), "actual": ly,
        "volume": round(ly * 0.98), "qtyPrice": round(ly * 0.92), "history": ly,
    }.get(method, ly)


def ai_json(base):
    lo = round(base * 0.9)
    hi = round(base * 1.05)
    mid = round((lo + hi) / 2)
    return json.dumps({"lo": lo, "hi": hi, "mid": mid,
                       "policy": "预算政策：规则基线", "basis": "往年预算", "exec": "上年执行约 95%"})


def comp_factor(code):
    h = int(hashlib.md5(code.encode("utf-8")).hexdigest(), 16)
    return 0.4 + (h % 90) / 100.0  # 0.4 ~ 1.3


# ---------------- 构建 organization（带前缀 code 防冲突） ----------------
orgs = [("HQ", "总部（上级部门）", None, "group")]
for bc, bn in business.items():
    orgs.append(("BU-" + bc, bn, "HQ", "business"))
for uc, un, bc in units:
    orgs.append(("U-" + uc, un, "BU-" + bc, "unit"))
for code, name in companies:
    orgs.append((code, name, "HQ", "company"))
for c in sorted(centers):
    orgs.append(("C-" + c, c, "HQ", "center"))
for dc, dn in depts:
    orgs.append(("D-" + dc, dn, "HQ", "dept"))


# ---------------- 写库 ----------------
conn = sqlite3.connect(DB)
cur = conn.cursor()
cur.execute("DELETE FROM organization")
cur.execute("DELETE FROM economic_event")
cur.execute("DELETE FROM unit_budget")

# 确保 account_subject 表与 economic_event.subject_id 列存在（独立运行导入脚本也安全）
cur.execute("""CREATE TABLE IF NOT EXISTS account_subject (
    id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT UNIQUE NOT NULL, name TEXT,
    cat TEXT, center TEXT, method TEXT, control_logic TEXT, parent_id INTEGER, sort_no INTEGER)""")
try:
    cur.execute("ALTER TABLE economic_event ADD COLUMN subject_id INTEGER")
except Exception:
    pass

org_id = {}
for code, name, parent, level in orgs:
    cur.execute("INSERT INTO organization (code,name,parent_id,level) VALUES (?,?,?,?)",
                (code, name, None, level))
    org_id[code] = cur.lastrowid
for code, name, parent, level in orgs:
    if parent and parent in org_id:
        cur.execute("UPDATE organization SET parent_id=? WHERE id=?", (org_id[parent], org_id[code]))

# 经济事项主数据（真实）；同时按 acct_code 去重建会计科目，并回填 subject_id
subject_rows = {}
sort_no = 0
for cat, info in events.items():
    # 编制方法不再由规则硬推：初始留空（待填），客户逻辑文字保留在 control_logic 作参考
    method = ""
    info["method"] = method
    acct = info["acct"]
    if acct and acct not in subject_rows:
        subject_rows[acct] = True
    last_year = det(cat, 500000, 5000000)
    base = baseline(method, last_year)
    amount = base
    monthly = decompose(amount)
    last_budget = round(last_year * 0.98)
    # 会计科目：先建（幂等，保留手动新增；已存在则刷新主数据字段）
    subj_id = None
    if acct:
        cur.execute("INSERT OR IGNORE INTO account_subject (code,name,cat,center,method,control_logic,sort_no) VALUES (?,?,?,?,?,?,?)",
                    (acct, cat, info["cat"] or None, info["center"] or None, method or None, logic_map.get(cat, ""), 0))
        if cur.rowcount == 0:
            cur.execute("UPDATE account_subject SET name=?, cat=?, center=?, method=?, control_logic=? WHERE code=?",
                        (cat, info["cat"] or None, info["center"] or None, method or None, logic_map.get(cat, ""), acct))
        subj_id = cur.execute("SELECT id FROM account_subject WHERE code=?", (acct,)).fetchone()[0]
    cur.execute(
        "INSERT INTO economic_event (cat,acct_code,center,amount,monthly,last_budget,last_year,method,ai,sort_no,subject_id) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (cat, acct, info["center"], amount, json.dumps(monthly), last_budget, last_year, method, ai_json(base), sort_no, subj_id))
    info["amount"] = amount
    info["last_year"] = last_year
    sort_no += 1

# 单位预算明细（生成式：经济事项基线 × 公司系数，关联真实组织）
for code, name in companies:
    cid = org_id[code]
    factor = comp_factor(code)
    for cat, info in events.items():
        amt = round(info["amount"] * factor)
        monthly = decompose(amt)
        cur.execute(
            "INSERT INTO unit_budget (org_id,cat,acct_code,amount,monthly,last_budget,last_year,method,ai) VALUES (?,?,?,?,?,?,?,?,?)",
            (cid, cat, info["acct"], amt, json.dumps(monthly),
             round(info["last_year"] * factor * 0.98), round(info["last_year"] * factor),
             info["method"], ai_json(amt)))

# demo 用户重映射（org code → 真实节点）
user_org = {
    "admin": "HQ", "zhangmy": "HQ", "xujing": "HQ",
    "lijing": "C-财务", "zhoufang": "C-总办", "sunyue": "BU-01",
    "chenkai": "C-总办", "liuyang": "3050", "wangmin": "2010",
    "zhaolei": "2010", "duanwei": "2020", "zhangwei": "2010",
}
for u, oc in user_org.items():
    if oc in org_id:
        cur.execute("UPDATE user SET org_id=? WHERE username=?", (org_id[oc], u))

conn.commit()
conn.close()
print("导入完成 → orgs=%d  events=%d  subjects=%d  unit_budget=%d  centers=%d  companies=%d" % (
    len(orgs), len(events), len(subject_rows), len(companies) * len(events), len(centers), len(companies)))
