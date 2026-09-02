#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract_subject_tree.py — 一次性脚本：从 2 个 Excel 抽取「经济事项 4 级分类树 + 叶子经济事项」，
产出 server/seeds/subject-tree.json（幂等 seed 数据源）。

权威口径：
  - 文件①「费控系统预算科目明细表」：B列「会计科目名称」按 '-' 拆分 = 4 级分类树（account_subject 节点）
    A列会计科目编码 / D列费用类型名称 = 叶子经济事项 / F列预算归口部门
  - 文件②「1.预算逻辑」：B列预算科目名称 -> F列 2026预算逻辑（写入叶子科目 control_logic）
"""
import openpyxl, collections, json, os

BASE = "/Users/yangjackson/AIProjects/BadgetManager"
XLSX = os.path.join(BASE, "docs/product/")
OUT = os.path.join(BASE, "server/seeds/subject-tree.json")
MAIN = "各中心费控系统2026年预算数据收集模板-财务部费控系统表单.xlsx"
LOGIC = "2.预算科目及逻辑-总经办表单.xlsx"


def load(fname):
    return openpyxl.load_workbook(os.path.join(XLSX, fname), data_only=True, read_only=True)


# ---------- 文件②：预算科目名称 -> 2026 预算逻辑 ----------
wb = load(LOGIC)
ws = wb["1.预算逻辑"]
logic_map = {}
for r in ws.iter_rows(values_only=True):
    if not r or not r[1]:
        continue
    name = str(r[1]).strip()
    if name == "预算科目名称":
        continue
    logic = str(r[5]).strip() if len(r) > 5 and r[5] else ""
    if name and name not in logic_map:
        logic_map[name] = logic
wb.close()

# ---------- 文件①：4 级分类树 + 叶子经济事项 ----------
wb = load(MAIN)
ws = wb["费控系统预算科目明细表"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
data = rows[1:]

# 树的节点（path 元组 -> 节点信息）；events 去重（D列 -> 叶子信息）
nodes = collections.OrderedDict()   # path tuple -> dict
events = collections.OrderedDict()  # D列 -> dict（去重）
node_first_seen = {}                # path tuple -> 首次出现的行序（用于 sortNo）

for idx, r in enumerate(data):
    b = str(r[1]).strip() if r[1] else ""
    d = str(r[3]).strip() if r[3] else ""
    a = str(r[0]).strip() if r[0] else ""
    e = str(r[4]).strip() if r[4] else ""
    f = str(r[5]).strip() if r[5] else ""
    c = str(r[2]).strip() if r[2] else ""
    # 客户数据里个别名称用 '/' 表示并列（如「生病住院/工伤慰问金」），与 path 分隔符冲突，统一换成顿号
    bp = tuple(x.strip().replace("/", "、") for x in b.split("-") if x.strip())
    if not bp:
        continue
    # 登记树节点（含所有前缀）
    for i in range(1, len(bp) + 1):
        prefix = bp[:i]
        if prefix not in nodes:
            nodes[prefix] = {
                "name": prefix[-1],
                "level": i,
                "path": "/".join(prefix),
                "cat": bp[0],
                "center": None,
                "controlLogic": "",
                "sortNo": None,
            }
            node_first_seen[prefix] = len(node_first_seen)
    # 叶子经济事项（D 列去重）
    if d and d not in events:
        events[d] = {
            "cat": d,
            "acctCode": a,
            "center": f,
            "subjectPath": list(bp),
            "method": "",
            "logic": logic_map.get(e, ""),
        }
    # 叶子科目节点：挂归口部门 + 控制逻辑（取首个非空，避免后行覆盖）
    leaf = nodes[bp]
    if f and not leaf["center"]:
        leaf["center"] = f
    if e in logic_map and not leaf["controlLogic"]:
        leaf["controlLogic"] = logic_map[e]

wb.close()

# 计算 sortNo：按「首次出现行序」稳定排序
order = sorted(nodes.keys(), key=lambda p: node_first_seen[p])
sort_no = 0
for p in order:
    nodes[p]["sortNo"] = sort_no
    sort_no += 1

subjects = [nodes[p] for p in order]

# parentPath：父节点路径（L1 无父）
for s in subjects:
    segs = s["path"].split("/")
    if len(segs) == 1:
        s["parentPath"] = None
    else:
        s["parentPath"] = "/".join(segs[:-1])

out = {
    "version": 1,
    "source": ["费控系统预算科目明细表(会计科目名称4级树)", "1.预算逻辑(控制逻辑)"],
    "subjects": subjects,
    "events": list(events.values()),
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as fp:
    json.dump(out, fp, ensure_ascii=False, indent=2)

# 统计
level_count = collections.Counter(s["level"] for s in subjects)
leaf_subjects = collections.Counter()
for e in out["events"]:
    leaf_subjects[len(e["subjectPath"])] += 1

print("抽取完成 →", OUT)
print("account_subject 树节点:", len(subjects), dict(sorted(level_count.items())))
print("  L1=%d  L2=%d  L3=%d  L4=%d" % (
    level_count.get(1, 0), level_count.get(2, 0), level_count.get(3, 0), level_count.get(4, 0)))
print("叶子经济事项(economic_event):", len(out["events"]), "个（按 subjectPath 深度分布:", dict(sorted(leaf_subjects.items())), "）")
print("带 controlLogic 的叶子科目:", sum(1 for s in subjects if s["controlLogic"]), "个")
